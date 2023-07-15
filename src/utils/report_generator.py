import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment


class ReportGenerator:

    @staticmethod
    def generate_excel_attendance(attendances):
        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active

        # Set up the column headers
        days = []
        for attendance in attendances:
            if attendance.date.weekday() < 5 and attendance.date not in days:
                days.append(attendance.date)
        days.sort()
        col = 2
        for day in days:
            sheet.cell(row=1, column=col, value=day.strftime("%d/%m/%Y"))
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")
            col += 1
        sheet.cell(row=1, column=1, value="Nombre completo")
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        sheet.cell(row=1, column=col, value="Total de asistencias")
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        # Write the attendance records
        row = 2
        for student in set(attendance.student for attendance in attendances):
            sheet.cell(row=row, column=1, value=f"{student.first_name} {student.last_name}")
            col = 2
            total_attendance = 0
            for day in days:
                attendance = next((a for a in attendances if a.student_id == student.id and a.date == day), None)
                if attendance is None:
                    sheet.cell(row=row, column=col, value="x")
                elif attendance.attended:
                    sheet.cell(row=row, column=col, value="✔")
                    total_attendance += 1
                else:
                    sheet.cell(row=row, column=col, value="P")
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                col += 1
            sheet.cell(row=row, column=col, value=total_attendance)
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            row += 1

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as file:
            workbook.save(file.name)
            return file.name
